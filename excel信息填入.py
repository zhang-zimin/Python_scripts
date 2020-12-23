import os
from openpyxl import load_workbook

path = r"D:\Data\帮忙\重合段单孔柱状图源文件"  # 更换为xlsx文件所在文件夹的目录
os.chdir(path)

files = os.listdir(path)


def hanshu(i, y):
    xunhuan = 1
    for a in files:
        if a != "path":
            # 提取
            workbook = load_workbook(filename=a)
            sheet = workbook["钻孔"]
            sheet = workbook.active
            danyuange = sheet[i].value
            # 输入
            workbook2 = load_workbook(filename="path")
            sheet2 = workbook2.active
            xunhuan += 1
            danyuange2 = sheet2.cell(y, xunhuan)
            danyuange2.value = danyuange
            # 保存
            workbook2.save(filename="path")


hanshu("B3", 1)  # B3是待提取文件的单元格编号，1是写入文件的行数
hanshu("B5", 2)
hanshu("B6", 3)
hanshu("B7", 4)
hanshu("B4", 5)
hanshu("B10", 6)
hanshu("B18", 7)
hanshu("B18", 8)

x = 10  # 岩性在写入从第几行开始写入
while x <= 57:
    for i in range(3, 27):
        hanshu("G" + str(i), x)
        hanshu("H" + str(i), x + 1)
        x += 2
