import openpyxl as vb

path1 = r"F:\code\1.xlsx"
path2 = r"F:\code\res6.xlsx"
工作簿 = vb.load_workbook(path1)
工作表 = 工作簿.active
res = vb.load_workbook(path2)
sheet = res.active
数据 = []
for j in range(1, 651, 11):
    区域 = 工作表["A" + str(j) + ":J" + str(j + 9)]
    for 行 in 区域:
        for 单元格 in 行:
            数据.append(单元格.value)
数据 = 数据[::-1]
for y in range(1, 61):
    for i in range(1, 101):
        sheet.cell(i, 61 - y).value = 数据[i + y * 100 - 101]
res.save(path2)
